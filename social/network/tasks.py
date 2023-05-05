from io import BytesIO
import subprocess
import requests

from django.conf import settings
from django.db import transaction
from django.utils import timezone
from django.core.files.base import File
from celery import Task
from celery import shared_task
from celery.utils.log import get_task_logger
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from twitter import tasks as twi_tasks
from linkedin import tasks as lin_tasks
from reusable.models import get_network_model


logger = get_task_logger(__name__)

NER_KEY_MAPPING = {
    "date": "تاریخ",
    "time": "زمان",
    "event": "رویداد",
    "money": "مالی",
    "person": "شخص",
    "percent": "درصد",
    "product": "محصول",
    "facility": "تسهیلات",
    "location": "مکان",
    "organization": "سازمان",
    "FAC": "سازه",
    "ORG": "سازمان",
    "GPE": "مکان",
    "LOC": "مکان",
    "EVENT": "رویداد",
    "MONEY": "مالی",
}


class BaseTaskWithRetry(Task):
    """A retry policy class for failed tasks

    Args:
        Task (func): a celery task
    """

    autoretry_for = (Exception,)
    retry_kwargs = {"max_retries": 10}
    retry_backoff = 5
    default_retry_delay = 60
    retry_jitter = True

    def run(self, *_args, **_kwargs):
        pass


@shared_task(base=BaseTaskWithRetry)
def extract_keywords(post_id):
    """We extract keywords for a post by using external service.
    We call an external api by post body
    We also consider ignored and blocked words
    We will delete blocked words
    We will ignore ignored words (we still save those into db)

    Args:
        post_id (int): id of the post
    """
    post_model = get_network_model("Post")
    channel_model = get_network_model("Channel")
    ignored_model = get_network_model("IgnoredKeyword")
    blocked_model = get_network_model("BlockedKeyword")
    keyword_model = get_network_model("Keyword")

    post = post_model.objects.get(id=post_id)
    ignored_keywords = list(ignored_model.objects.values_list("keyword", flat=True))
    blocked_keywords = list(blocked_model.objects.values_list("keyword", flat=True))

    endpoint = None
    if post.channel.language == channel_model.PERSIAN:
        endpoint = "http://persian_analyzer_api/v1/app/keyword/"
    if post.channel.language == channel_model.ENGLISH:
        endpoint = "http://analyzer_api/api/v1/keyword/"

    resp = requests.post(endpoint, {"text": post.body}, timeout=5).json()
    objs = []
    words = resp["keywords"]
    if "keyphrases" in resp:
        words += resp["keyphrases"]
    for keyword in words:
        if keyword not in blocked_keywords:
            ignored = keyword in ignored_keywords
            objs.append(keyword_model(post=post, keyword=keyword, ignored=ignored))

    keyword_model.objects.bulk_create(objs)


@shared_task(base=BaseTaskWithRetry)
def extract_ner(post_id):
    """We extract net entities by using an external service.
    We call an external api and send post's body to it and then
    Store result to db.

    Args:
        post_id (int): id of the post
    """
    post_model = get_network_model("Post")
    channel_model = get_network_model("Channel")

    with transaction.atomic():
        post = post_model.objects.select_for_update().get(id=post_id)
        endpoint = None
        if post.channel.language == channel_model.PERSIAN:
            endpoint = "http://persian_analyzer_api/v1/app/ner/"
        if post.channel.language == channel_model.ENGLISH:
            endpoint = "http://analyzer_api/api/v1/ner/"
        resp = requests.post(endpoint, {"text": post.body}, timeout=5).json()
        temp = {}
        for key, _ in NER_KEY_MAPPING.items():
            if key in resp:
                temp[NER_KEY_MAPPING[key]] = list(set(resp.pop(key)))
        post.ner = temp
        post.save()


@shared_task(base=BaseTaskWithRetry)
def extract_sentiment(post_id):
    """We get sentiment by using external api.
    We send body of a post to the api

    Args:
        post_id (int): id of the post
    """
    post_model = get_network_model("Post")

    with transaction.atomic():
        post = post_model.objects.select_for_update().get(id=post_id)
        resp = requests.post(
            "http://persian_analyzer_api/v1/app/sentiment/",
            {"text": post.body},
            timeout=5,
        ).json()
        post.sentiment = resp
        post.save()


@shared_task(base=BaseTaskWithRetry)
def extract_categories(post_id):
    """We get categories of a post by using external api.
    We send the body of a post to the api and get result in shape of category_a: 20%

    Args:
        post_id (int): id of the post
    """
    post_model = get_network_model("Post")

    with transaction.atomic():
        post = post_model.objects.select_for_update().get(id=post_id)
        resp = requests.post(
            "http://persian_analyzer_api/v1/app/classification/",
            {"text": post.body},
            timeout=5,
        ).json()
        sorted_result = sorted(resp, key=lambda k: k["score"], reverse=True)
        post.category = sorted_result
        post.main_category_title = sorted_result[0]["label"]
        post.save()


@shared_task()
def check_channels_crawl():
    """In this task we check which channels we should crawl at time of running the task
    This is a periodic task.
    """
    channel_model = get_network_model("Channel")

    channels = channel_model.objects.filter(last_crawl__isnull=False)
    for channel in channels:
        interval = timezone.localtime() - channel.last_crawl
        hours = interval.total_seconds() / 3600
        if hours >= channel.crawl_interval:
            print(f"******* channel {channel} must crawled")
            if channel.network.name == "Twitter":
                twi_tasks.get_twitter_posts(channel.pk)
            elif channel.network.name == "Linkedin":
                lin_tasks.get_linkedin_posts(channel.pk)


@shared_task()
def take_backup(backup_id):
    """This task creates a backup from our postgres database.
    This backup will be created in a sql.gz file type and then will
    be saved in the filesystem.

    Args:
        backup_id (int): id of backup row in the administration panel.
        (Admin create a row and then we run this task)
    """
    backup_model = get_network_model("Backup")

    backup = backup_model.objects.get(pk=backup_id)
    date_time = backup.created_at.strftime("%d-%m-%Y-%H_%M_%S")

    if backup.type == backup_model.RASAD_1:
        subprocess.run(
            [
                "ssh",
                "-i",
                "/app/secrets/id_rsa_social_api",
                "-o",
                "StrictHostKeyChecking=no",
                f"root@{settings.SERVER_IP}",
                f"docker exec -t postgres pg_dumpall -c -U postgres \
                    | gzip > /root/army/frontend/dist/backup/postgres_db_{date_time}.sql.gz",
            ],
            check=False,
        )
        backup.link = (
            f"http://{settings.SERVER_IP}/backup/postgres_db_{date_time}.sql.gz"
        )
    elif backup.type == backup_model.RASAD_2:
        subprocess.run(
            [
                "ssh",
                "-i",
                "/app/secrets/id_rsa_social_api",
                "-o",
                "StrictHostKeyChecking=no",
                f"root@{settings.SERVER_IP}",
                f"docker exec -t social_db pg_dumpall -c -U postgres | \
                    gzip > /root/army/frontend/dist/backup/social_db_{date_time}.sql.gz",
            ],
            check=False,
        )
        backup.link = f"http://{settings.SERVER_IP}/backup/social_db_{date_time}.sql.gz"

    backup.status = backup_model.COMPLETED
    backup.save()


@shared_task()
def export_channel_list(export_id):
    """This function create a list of channel in a excel file.
    User can download that file.

    Args:
        export_id (int): This is the id of the report.
        (Admin first create report row. then we run this task.)
    """
    channel_model = get_network_model("Channel")
    channel_list_export_model = get_network_model("ChannelListExport")

    channels = channel_model.objects.all()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.freeze_panes = worksheet["B2"]
    worksheet.title = "Channel List Report"
    fields_name = ["Number", "Name", "Network", "Status"]

    for col_num, column_title in enumerate(fields_name, start=1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=Side(border_style="medium", color="FF000000"))
        cell.value = column_title

    for row_index, channel in enumerate(channels, start=2):
        guest_fields = [
            row_index - 1,
            channel.name,
            channel.network.name,
            channel.status,
        ]
        for col_index, column in enumerate(guest_fields, start=1):
            cell = worksheet.cell(row=row_index, column=col_index)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.value = column

    virtual_workbook = BytesIO()
    workbook.save(virtual_workbook)
    export = channel_list_export_model.objects.get(pk=export_id)
    export.file.save("excel.xlsx", File(virtual_workbook))
    export.save()


@shared_task()
def remove_blocked_keywords():
    """This function retrieve list of block words then
    by using them, delete blocked keywords.
    """
    blocked_model = get_network_model("BlockedKeyword")
    keyword_model = get_network_model("Keyword")

    blocked_keywords = list(blocked_model.objects.values_list("keyword", flat=True))

    # for performance consideration, we loop over them in batch mode
    first_id = keyword_model.objects.first().id
    last_id = keyword_model.objects.last().id
    batch_size = 10000
    current_counter = first_id

    while current_counter < last_id:
        for item in keyword_model.objects.filter(
            id__gte=current_counter, id__lte=current_counter + batch_size
        ):
            if item.keyword in blocked_keywords:
                item.delete()
        current_counter += batch_size


@shared_task()
def test_error():
    """This is a test function
    We use it to check if our logger works correctly.
    """
    logger.error("error")
